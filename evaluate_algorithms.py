import time
import pandas as pd
import sys
sys.stdout.reconfigure(encoding='utf-8')
import os
import random

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from src.core.level import Level
from src.algorithms.solver_adapter import SolverAdapter

from src.algorithms.basic_search.simple_hill_climbing import simple_hill_climbing
from src.algorithms.basic_search.steepest_ascent import steepest_ascent_hill_climbing
from src.algorithms.stochastic_search.stochastic_hc import stochastic_hill_climbing
from src.algorithms.stochastic_search.first_choice_hc import first_choice_hill_climbing
from src.algorithms.escape_maxima.backtracking_hc import backtracking_hill_climbing
from src.algorithms.escape_maxima.jumping_hc import jumping_hill_climbing
from src.algorithms.escape_maxima.random_restart_hc import random_restart_hill_climbing


def make_jump_function(adapter):
    """
    Tạo hàm nhảy ngẫu nhiên cho Jumping HC.
    Thực hiện 5-15 bước di chuyển hợp lệ ngẫu nhiên từ trạng thái hiện tại
    để thoát khỏi Local Minima.
    """
    def generate_random_jump(current_state):
        import random
        state = current_state
        actions = []
        num_jumps = random.randint(5, 15)
        
        for _ in range(num_jumps):
            neighbors = adapter.get_neighbors(state)
            if not neighbors:
                break
            action, next_state = random.choice(neighbors)
            state = next_state
            actions.append(action)
        
        if actions:
            return state, actions
        return None, []
    
    return generate_random_jump

def make_restart_function(initial_state, adapter):
    """
    Tạo hàm khởi động lại cho Random Restart HC.
    Đưa nhân vật đi ngẫu nhiên 3-10 bước từ trạng thái ban đầu
    để tạo điểm xuất phát mới mỗi lần restart.
    """
    def generate_random_initial():
        import random
        state = initial_state
        num_moves = random.randint(3, 10)
        
        for _ in range(num_moves):
            neighbors = adapter.get_neighbors(state)
            if not neighbors:
                break
            _, next_state = random.choice(neighbors)
            state = next_state
        
        return state
    
    return generate_random_initial

def trace_heuristic_history(initial_state, path, solver_adapter):
    """
    Takes the final taken path and reconstructs the heuristic states step-by-step
    to show exactly how the algorithm navigated down the heuristic landscape.
    """
    get_neighbors = solver_adapter.get_neighbors
    original_h = solver_adapter.get_heuristic_func()
    
    history = []
    current_state = initial_state
    
    history.append(original_h(current_state))
    
    for action_taken in path:
        neighbors = get_neighbors(current_state)
        found = False
        for a, s in neighbors:
            if a == action_taken:
                current_state = s
                found = True
                break
        # Just in case our fast 'get_neighbors' doesn't align with jump logic, fallback handle
        if not found:
            pass # Ideally path is strictly valid neighbors
        
        history.append(original_h(current_state))
        
    return history

def run_evaluation(target_level=None):
    if target_level is None:
        print("Đang tải bàn cờ mặc định (Map)...")
        level = Level()
        # Chạy trên map mặc định, bạn có thể đổi tên tùy ý
        level.load_from_file("src/map/map_default.xlsx")
    else:
        print("Đang lấy mốc bàn cờ hiện tại từ game...")
        level = target_level
    
    adapter = SolverAdapter(level)
    initial_state = adapter.get_initial_state()
    get_neighbors = adapter.get_neighbors
    get_heuristic = adapter.get_heuristic_func()
    
    # Tạo hàm hỗ trợ thật cho Jumping HC và Random Restart HC
    jump_fn = make_jump_function(adapter)
    restart_fn = make_restart_function(initial_state, adapter)
    
    algorithms = [
        ("Simple Hill Climbing",      simple_hill_climbing,          [5000]),
        ("Steepest Ascent",           steepest_ascent_hill_climbing,  [5000]),
        ("Stochastic Hill Climbing",  stochastic_hill_climbing,       [5000]),
        ("First Choice Hill Climbing",first_choice_hill_climbing,     [5000]),
        ("Backtracking Hill Climbing",backtracking_hill_climbing,     [5000]),
        ("Jumping Hill Climbing",     jumping_hill_climbing,          [jump_fn, 5000]),
        ("Random Restart HC",         random_restart_hill_climbing,   [restart_fn, 10, 500]),
    ]
    
    results = []
    max_steps = 0 # To align excel columns nicely
    
    print("Bắt đầu chạy thí nghiệm 7 thuật toán...\n")
    for name, func, extra_args in algorithms:
        print(f"-> Đang chạy: {name} ...", end=" ")
        start_time = time.time()
        
        # Gọi hàm
        best_state, path = func(initial_state, get_neighbors, get_heuristic, *extra_args)
        
        end_time = time.time()
        exec_time_ms = (end_time - start_time) * 1000
        
        print(f"Hoàn thành ({exec_time_ms:.2f} ms). H(cuối)={get_heuristic(best_state)}")
        
        # Truy vết lịch sử Heuristic (đường đi hàm giảm)
        h_history = trace_heuristic_history(initial_state, path, adapter)
        max_steps = max(max_steps, len(h_history))
        
        results.append({
            "Thuật toán": name,
            "Tốc độ (ms)": round(exec_time_ms, 2),
            "Số bước đi": len(path),
            "H(Start)": h_history[0],
            "H(End)": get_heuristic(best_state),
            "h_history_list": h_history,
            "path": path
        })
        
    print("\nQuá trình chạy hoàn tất. Đang cập nhật ra file dạng bảng tách biệt...")
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter

    filename = "Kq_Thuật_toán_AI_Bảng.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kq AI"

    # Định nghĩa màu sắc (Hex codes)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Xanh dương
    header_font = Font(color="FFFFFF", bold=True)
    algo_font = Font(color="1F497D", bold=True) # Xanh đậm
    time_font = Font(color="00B050", bold=True) # Xanh lá cây
    center_align = Alignment(horizontal='center', vertical='center')

    # Dòng 1 & 2: Tên thuật toán & Tốc độ
    # Dòng 3: Tiêu đề cột chi tiết
    for c_idx, r in enumerate(results):
        base_col = c_idx * 3 + 1
        
        # Algo Name
        ws.cell(row=1, column=base_col+1, value=f"Thuật toán: {r['Thuật toán']}")
        ws.cell(row=1, column=base_col+1).font = algo_font
        ws.cell(row=1, column=base_col+1).alignment = center_align
        
        # Time
        ws.cell(row=2, column=base_col+1, value=f"Thời gian: {r['Tốc độ (ms)']} ms")
        ws.cell(row=2, column=base_col+1).font = time_font
        ws.cell(row=2, column=base_col+1).alignment = center_align

        # Headers
        headers = ["Bước thứ", "Hành động", "Heuristic"]
        for j, h in enumerate(headers):
            c = ws.cell(row=3, column=base_col+j, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center_align

    # Lấy số vòng lặp tối đa
    max_steps = max([len(r["path"]) for r in results]) if results else 0
    
    # Dòng 4 trở đi: Kết quả từng bước
    for step in range(-1, max_steps):
        row_idx = step + 5 # Bắt đầu in từ dòng thứ 4 (vì -1 + 5 = 4)
        for c_idx, r in enumerate(results):
            base_col = c_idx * 3 + 1
            path_list = r["path"]
            h_list = r["h_history_list"]
            
            if step == -1:
                ws.cell(row=row_idx, column=base_col+0, value=0).alignment = center_align
                ws.cell(row=row_idx, column=base_col+1, value="Khởi đầu").alignment = center_align
                ws.cell(row=row_idx, column=base_col+2, value=h_list[0]).alignment = center_align
            else:
                if step < len(path_list):
                    action = path_list[step]
                    h_val = h_list[step+1] if step+1 < len(h_list) else "N/A"
                    ws.cell(row=row_idx, column=base_col+0, value=step+1).alignment = center_align
                    ws.cell(row=row_idx, column=base_col+1, value=action).alignment = center_align
                    ws.cell(row=row_idx, column=base_col+2, value=h_val).alignment = center_align

    # Phần Kết Luận ở cuối bảng
    conclusion_row = max_steps + 7
    for c_idx, r in enumerate(results):
        base_col = c_idx * 3 + 1
        name = r["Thuật toán"]
        h_end = r["H(End)"]
        steps = r["Số bước đi"]
        time_ms = r["Tốc độ (ms)"]
        
        if h_end == 0:
            conclusion = f"✓ Thành công: Phá đảo trong {steps} bước ({time_ms} ms)."
        else:
            conclusion = f"✗ Thất bại: Bị kẹt tại đỉnh cục bộ \n(H = {h_end})."
            
        ws.merge_cells(start_row=conclusion_row, start_column=base_col, end_row=conclusion_row+1, end_column=base_col+2)
        cell = ws.cell(row=conclusion_row, column=base_col, value=conclusion)
        cell.font = Font(bold=True, color="9C0006" if h_end > 0 else "006100") # Xanh nếu thắng, Đỏ sẫm nếu thua
        cell.fill = PatternFill(start_color="FFC7CE" if h_end > 0 else "C6EFCE", end_color="FFC7CE" if h_end > 0 else "C6EFCE", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Chỉnh lại cỡ cột cho đẹp
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2: BIỂU ĐỒ ĐÁNH GIÁ TỐC ĐỘ CÁC THUẬT TOÁN
    # ══════════════════════════════════════════════════════════════════════════
    print("Đang vẽ biểu đồ đánh giá tốc độ vào file Excel...")
    
    ws_chart = wb.create_sheet(title="Biểu Đồ Tốc Độ")
    
    # ── Bảng dữ liệu tổng hợp cho biểu đồ (ẩn phía trên) ──────────────────
    # Màu sắc riêng cho từng thuật toán (dùng cho DataPoint)
    ALGO_CHART_COLORS = [
        "4E91D2",  # Simple HC - Xanh dương
        "E07B39",  # Steepest Ascent - Cam
        "5BAD72",  # Stochastic HC - Xanh lá
        "C94F7C",  # First Choice HC - Hồng
        "8A63D2",  # Backtracking HC - Tím
        "D13B3B",  # Jumping HC - Đỏ
        "3BBBD4",  # Random Restart HC - Xanh ngọc
    ]
    
    title_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    subtitle_font = Font(color="FFFFFF", bold=True, size=11)
    data_header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    data_header_font = Font(color="E5E7EB", bold=True, size=10)
    data_font = Font(color="D1D5DB", size=10)
    data_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='374151'),
        right=Side(style='thin', color='374151'),
        top=Side(style='thin', color='374151'),
        bottom=Side(style='thin', color='374151'),
    )
    
    # Tiêu đề lớn
    ws_chart.merge_cells("A1:N1")
    title_cell = ws_chart.cell(row=1, column=1, value="📊 BIỂU ĐỒ ĐÁNH GIÁ TỐC ĐỘ CÁC THUẬT TOÁN HILL CLIMBING – SOKOBAN AI")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_chart.row_dimensions[1].height = 40
    
    # ── Bảng 1: Dữ liệu tổng hợp (dòng 3-11) ─────────────────────────────
    summary_headers = ["Thuật toán", "Tốc độ (ms)", "Số bước đi", "H(Start)", "H(End)", 
                       "Cải thiện H", "Kết quả"]
    
    ws_chart.merge_cells("A3:G3")
    sub_cell = ws_chart.cell(row=3, column=1, value="📋 BẢNG DỮ LIỆU TỔNG HỢP")
    sub_cell.font = subtitle_font
    sub_cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_chart.row_dimensions[3].height = 28
    
    for j, header in enumerate(summary_headers):
        c = ws_chart.cell(row=4, column=j+1, value=header)
        c.font = data_header_font
        c.fill = data_header_fill
        c.alignment = center_align
        c.border = thin_border
    
    for i, r in enumerate(results):
        row = 5 + i
        h_improve = r["H(Start)"] - r["H(End)"]
        kq = "✓ Thành công" if r["H(End)"] == 0 else f"✗ Thất bại (H={r['H(End)']})"
        
        values = [
            r["Thuật toán"],
            r["Tốc độ (ms)"],
            r["Số bước đi"],
            r["H(Start)"],
            r["H(End)"],
            h_improve,
            kq
        ]
        for j, val in enumerate(values):
            c = ws_chart.cell(row=row, column=j+1, value=val)
            c.font = data_font
            c.fill = data_fill
            c.alignment = center_align
            c.border = thin_border
            # Tô màu kết quả
            if j == 6:
                if r["H(End)"] == 0:
                    c.font = Font(color="22C55E", bold=True, size=10)
                else:
                    c.font = Font(color="EF4444", bold=True, size=10)
    
    # Chỉnh cỡ cột bảng tổng hợp
    col_widths_summary = [30, 15, 15, 12, 12, 15, 25]
    for j, w in enumerate(col_widths_summary):
        ws_chart.column_dimensions[get_column_letter(j+1)].width = w
    
    # ── BIỂU ĐỒ 1: Tốc độ thực thi (Bar Chart) ────────────────────────────
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "⏱ Thời Gian Thực Thi (ms)"
    chart1.y_axis.title = "Thời gian (ms)"
    chart1.x_axis.title = "Thuật toán"
    chart1.width = 28
    chart1.height = 16
    
    # Dữ liệu: cột B (Tốc độ), dòng 4-11
    data1 = Reference(ws_chart, min_col=2, min_row=4, max_row=4+len(results))
    cats1 = Reference(ws_chart, min_col=1, min_row=5, max_row=4+len(results))
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.shape = 4
    
    # Tô màu riêng từng cột
    series1 = chart1.series[0]
    series1.graphicalProperties.line.noFill = True
    for i in range(len(results)):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = ALGO_CHART_COLORS[i % len(ALGO_CHART_COLORS)]
        series1.data_points.append(pt)
    
    # Hiện giá trị trên đầu cột
    series1.dLbls = DataLabelList()
    series1.dLbls.showVal = True
    series1.dLbls.numFmt = '0.00'
    
    chart1.legend = None
    ws_chart.add_chart(chart1, "A14")
    
    # ── BIỂU ĐỒ 2: Số bước đi (Bar Chart) ─────────────────────────────────
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "👣 Số Bước Đi"
    chart2.y_axis.title = "Số bước"
    chart2.x_axis.title = "Thuật toán"
    chart2.width = 28
    chart2.height = 16
    
    data2 = Reference(ws_chart, min_col=3, min_row=4, max_row=4+len(results))
    cats2 = Reference(ws_chart, min_col=1, min_row=5, max_row=4+len(results))
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.shape = 4
    
    series2 = chart2.series[0]
    series2.graphicalProperties.line.noFill = True
    for i in range(len(results)):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = ALGO_CHART_COLORS[i % len(ALGO_CHART_COLORS)]
        series2.data_points.append(pt)
    
    series2.dLbls = DataLabelList()
    series2.dLbls.showVal = True
    series2.dLbls.numFmt = '0'
    
    chart2.legend = None
    ws_chart.add_chart(chart2, "A31")
    
    # ── BIỂU ĐỒ 3: Mức cải thiện Heuristic (Bar Chart ngang) ──────────────
    chart3 = BarChart()
    chart3.type = "bar"  # horizontal
    chart3.style = 10
    chart3.title = "📉 Mức Cải Thiện Heuristic (H_start − H_end)"
    chart3.x_axis.title = "Cải thiện H"
    chart3.y_axis.title = "Thuật toán"
    chart3.width = 28
    chart3.height = 16
    
    data3 = Reference(ws_chart, min_col=6, min_row=4, max_row=4+len(results))
    cats3 = Reference(ws_chart, min_col=1, min_row=5, max_row=4+len(results))
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.shape = 4
    
    series3 = chart3.series[0]
    series3.graphicalProperties.line.noFill = True
    for i in range(len(results)):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = ALGO_CHART_COLORS[i % len(ALGO_CHART_COLORS)]
        series3.data_points.append(pt)
    
    series3.dLbls = DataLabelList()
    series3.dLbls.showVal = True
    series3.dLbls.numFmt = '0'
    
    chart3.legend = None
    ws_chart.add_chart(chart3, "A48")
    
    # ── BIỂU ĐỒ 4: Đường hội tụ Heuristic (Line Chart) ────────────────────
    # Tạo bảng dữ liệu riêng cho biểu đồ đường (dòng từ max_convergence_row)
    conv_start_row = 4 + len(results) + 3  # Sau bảng tổng hợp, cách 2 dòng
    
    ws_chart.merge_cells(start_row=conv_start_row, start_column=9, 
                         end_row=conv_start_row, end_column=9+len(results))
    conv_title = ws_chart.cell(row=conv_start_row, column=9, 
                               value="📈 DỮ LIỆU HỘI TỤ HEURISTIC")
    conv_title.font = subtitle_font
    conv_title.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    conv_title.alignment = Alignment(horizontal='center', vertical='center')
    
    # Header: "Bước" + tên 7 thuật toán
    header_row = conv_start_row + 1
    c = ws_chart.cell(row=header_row, column=9, value="Bước")
    c.font = data_header_font
    c.fill = data_header_fill
    c.alignment = center_align
    c.border = thin_border
    
    for i, r in enumerate(results):
        c = ws_chart.cell(row=header_row, column=10+i, value=r["Thuật toán"])
        c.font = data_header_font
        c.fill = data_header_fill
        c.alignment = center_align
        c.border = thin_border
        ws_chart.column_dimensions[get_column_letter(10+i)].width = 18
    
    ws_chart.column_dimensions[get_column_letter(9)].width = 10
    
    # Chuẩn hóa tiến trình → lấy mẫu 50 điểm để biểu đồ không quá nặng
    max_h_len = max(len(r["h_history_list"]) for r in results)
    num_sample_points = min(50, max_h_len)
    
    for pt_idx in range(num_sample_points):
        data_row = header_row + 1 + pt_idx
        # Cột bước (% tiến trình)
        progress_pct = round(pt_idx / max(num_sample_points - 1, 1) * 100)
        c = ws_chart.cell(row=data_row, column=9, value=progress_pct)
        c.font = data_font
        c.fill = data_fill
        c.alignment = center_align
        c.border = thin_border
        
        for algo_idx, r in enumerate(results):
            h_list = r["h_history_list"]
            # Nội suy lấy giá trị tại vị trí tương ứng
            float_idx = pt_idx / max(num_sample_points - 1, 1) * max(len(h_list) - 1, 0)
            lo = int(float_idx)
            hi = min(lo + 1, len(h_list) - 1)
            frac = float_idx - lo
            if lo < len(h_list) and hi < len(h_list):
                h_val = h_list[lo] * (1 - frac) + h_list[hi] * frac
            elif h_list:
                h_val = h_list[-1]
            else:
                h_val = 0
            
            c = ws_chart.cell(row=data_row, column=10+algo_idx, value=round(h_val, 1))
            c.font = data_font
            c.fill = data_fill
            c.alignment = center_align
            c.border = thin_border
    
    # Vẽ biểu đồ đường
    chart4 = LineChart()
    chart4.title = "📈 Đường Hội Tụ Heuristic Theo Tiến Trình Tìm Kiếm"
    chart4.x_axis.title = "Tiến trình (%)"
    chart4.y_axis.title = "Giá trị Heuristic H"
    chart4.width = 36
    chart4.height = 20
    chart4.style = 10
    
    data4 = Reference(ws_chart, min_col=10, max_col=9+len(results), 
                       min_row=header_row, max_row=header_row+num_sample_points)
    cats4 = Reference(ws_chart, min_col=9, min_row=header_row+1, 
                       max_row=header_row+num_sample_points)
    chart4.add_data(data4, titles_from_data=True)
    chart4.set_categories(cats4)
    
    # Tô màu cho từng đường
    LINE_COLORS = ["4E91D2", "E07B39", "5BAD72", "C94F7C", "8A63D2", "D13B3B", "3BBBD4"]
    for i, s in enumerate(chart4.series):
        s.graphicalProperties.line.solidFill = LINE_COLORS[i % len(LINE_COLORS)]
        s.graphicalProperties.line.width = 22000  # ~2pt
        s.smooth = True
    
    ws_chart.add_chart(chart4, "A65")
    
    # ── Tô nền cho toàn sheet biểu đồ ──────────────────────────────────────
    bg_fill = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
    for row in ws_chart.iter_rows(min_row=1, max_row=100, min_col=1, max_col=20):
        for cell in row:
            if cell.fill == PatternFill() or cell.fill.start_color.index == '00000000':
                cell.fill = bg_fill

    wb.save(filename)
    print(f"\n[THÀNH CÔNG] Đã lưu bảng kết quả + biểu đồ đánh giá tốc độ vào: {filename}")
    print(f"   → Sheet 'Kq AI': Bảng chi tiết từng bước")
    print(f"   → Sheet 'Biểu Đồ Tốc Độ': 4 biểu đồ đánh giá trực quan")

    # --- Tìm đường đi tốt nhất để trả về ---
    best_algo = None
    best_h = float('inf')
    best_path = []
    
    for r in results:
        h_end = r["H(End)"]
        steps = r["Số bước đi"]
        
        # Tiêu chí: H thấp nhất, nếu bằng thì ưu tiên số bước ít hơn
        if h_end < best_h or (h_end == best_h and steps < len(best_path)):
            best_h = h_end
            best_path = r["path"]
            best_algo = r["Thuật toán"]
            
    print(f"\n[🚀 AI AUTO-DRIVE] Thuật toán ưu việt nhất: {best_algo} | H={best_h} ({len(best_path)} bước)")
    
    return best_path

if __name__ == "__main__":
    run_evaluation()
