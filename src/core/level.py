import os
from src.utils.constants import WALL, TARGET, BOX, PLAYER, BOX_ON_TARGET, PLAYER_ON_TARGET, EMPTY, OUTSIDE
from src.core.grid import Grid
from src.entities.player import Player
from src.entities.box import Box

class Level:
    def __init__(self):
        self.grid = Grid()
        self.player = None
        self.boxes = []
        self.width = 0
        self.height = 0

    def load_from_file(self, filepath):
        if not os.path.exists(filepath):
            print(f"Level file not found: {filepath}")
            return False

        if filepath.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, data_only=True)
                sheet = wb.active
                lines = []
                max_col = sheet.max_column
                for row in range(1, sheet.max_row + 1):
                    line_chars = []
                    for col in range(1, max_col + 1):
                        val = sheet.cell(row=row, column=col).value
                        if val is None:
                            line_chars.append(' ')
                        else:
                            text = str(val).strip()
                            char = text[0] if text else ' '
                            line_chars.append(char)
                    lines.append(''.join(line_chars).rstrip())
                return self.load_from_lines(lines)
            except Exception as e:
                print(f"[LỖI] Không thể đọc map từ Excel: {e}")
                return False
                
        with open(filepath, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        return self.load_from_lines(lines)

    def load_from_lines(self, lines):
        self.boxes = []
        self.player = None
        
        cleaned_lines = [line.rstrip('\n\r') for line in lines]
        
        # Xoá các dòng hoàn toàn rỗng ở cuối (do Excel hay quét dư max_row)
        while cleaned_lines and not cleaned_lines[-1].strip():
            cleaned_lines.pop()
            
        self.height = len(cleaned_lines)
        if self.height == 0:
            return False
            
        self.width = max(len(line) for line in cleaned_lines)
        
        # Sửa lỗi vẽ sàn tràn lan: Default pad bằng OUTSIDE thay vì EMPTY
        grid_data = [[OUTSIDE for _ in range(self.width)] for _ in range(self.height)]

        for y, line in enumerate(cleaned_lines):
            for x, char in enumerate(line):
                if char == WALL:
                    grid_data[y][x] = WALL
                elif char == TARGET:
                    grid_data[y][x] = TARGET
                elif char == BOX:
                    grid_data[y][x] = EMPTY
                    self.boxes.append(Box(x, y))
                elif char == PLAYER:
                    grid_data[y][x] = EMPTY
                    self.player = Player(x, y)
                elif char == BOX_ON_TARGET:
                    grid_data[y][x] = TARGET
                    box = Box(x, y)
                    box.is_on_target = True
                    self.boxes.append(box)
                elif char == PLAYER_ON_TARGET:
                    grid_data[y][x] = TARGET
                    self.player = Player(x, y)
                elif char == OUTSIDE:
                    grid_data[y][x] = OUTSIDE
                elif char == EMPTY:
                    grid_data[y][x] = EMPTY
                else:
                    grid_data[y][x] = EMPTY

        self.grid.load_from_2d_array(grid_data)
        
        if self.player is None:
            print("Warning: No player found in level")
            return False
            
        return True
