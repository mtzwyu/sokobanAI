"""
chart_analysis.py
=================
Script phân tích và vẽ biểu đồ đánh giá tốc độ / hiệu suất
của 7 thuật toán Hill Climbing trong bài toán Sokoban.

Cách dùng:
    python chart_analysis.py           # Chạy tất cả thuật toán rồi vẽ biểu đồ
    python chart_analysis.py --excel   # Đọc file Excel đã có rồi vẽ biểu đồ

Kết quả: file "Bieu_Do_Danh_Gia_Thuat_Toan.png" được lưu cùng thư mục gốc.
"""

import sys
import os
import time
import random
import argparse

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
sys.stdout.reconfigure(encoding="utf-8")

import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
from adjustText import adjust_text

matplotlib.rcParams["font.family"] = "Segoe UI"
matplotlib.rcParams["axes.unicode_minus"] = False

# ──────────────────────────────────────────────────────────────────────────────
# Bảng màu cố định cho từng thuật toán (nhất quán mọi biểu đồ)
# ──────────────────────────────────────────────────────────────────────────────
ALGO_COLORS = {
    "Simple Hill Climbing":       "#4E91D2",
    "Steepest Ascent":            "#E07B39",
    "Stochastic Hill Climbing":   "#5BAD72",
    "First Choice Hill Climbing": "#C94F7C",
    "Backtracking Hill Climbing": "#8A63D2",
    "Jumping Hill Climbing":      "#D13B3B",
    "Random Restart HC":          "#3BBBD4",
}

OUTPUT_PNG  = "Bieu_Do_Danh_Gia_Thuat_Toan.png"
EXCEL_INPUT = "Kq_Thuật_toán_AI_Bảng.xlsx"


# ══════════════════════════════════════════════════════════════════════════════
# 1.  THU THẬP DỮ LIỆU
# ══════════════════════════════════════════════════════════════════════════════

def _make_jump_fn(adapter):
    def jump(state):
        s = state
        for _ in range(random.randint(5, 15)):
            nb = adapter.get_neighbors(s)
            if not nb:
                break
            _, s = random.choice(nb)
        return s, []
    return jump

def _make_restart_fn(initial_state, adapter):
    def restart():
        s = initial_state
        for _ in range(random.randint(3, 10)):
            nb = adapter.get_neighbors(s)
            if not nb:
                break
            _, s = random.choice(nb)
        return s
    return restart

def _trace_h_history(initial_state, path, adapter):
    h = adapter.get_heuristic_func()
    get_nb = adapter.get_neighbors
    history = [h(initial_state)]
    cur = initial_state
    for action in path:
        for a, s in get_nb(cur):
            if a == action:
                cur = s
                break
        history.append(h(cur))
    return history


def collect_from_run():
    """Chạy lại tất cả thuật toán và trả về list dict kết quả."""
    from src.core.level import Level
    from src.algorithms.solver_adapter import SolverAdapter
    from src.algorithms.basic_search.simple_hill_climbing      import simple_hill_climbing
    from src.algorithms.basic_search.steepest_ascent           import steepest_ascent_hill_climbing
    from src.algorithms.stochastic_search.stochastic_hc        import stochastic_hill_climbing
    from src.algorithms.stochastic_search.first_choice_hc      import first_choice_hill_climbing
    from src.algorithms.escape_maxima.backtracking_hc          import backtracking_hill_climbing
    from src.algorithms.escape_maxima.jumping_hc               import jumping_hill_climbing
    from src.algorithms.escape_maxima.random_restart_hc        import random_restart_hill_climbing

    level = Level()
    level.load_from_file("src/map/map_default.xlsx")
    adapter = SolverAdapter(level)
    ini     = adapter.get_initial_state()
    get_nb  = adapter.get_neighbors
    get_h   = adapter.get_heuristic_func()

    jump_fn    = _make_jump_fn(adapter)
    restart_fn = _make_restart_fn(ini, adapter)

    ALGOS = [
        ("Simple Hill Climbing",       simple_hill_climbing,          [5000]),
        ("Steepest Ascent",            steepest_ascent_hill_climbing,  [5000]),
        ("Stochastic Hill Climbing",   stochastic_hill_climbing,       [5000]),
        ("First Choice Hill Climbing", first_choice_hill_climbing,     [5000]),
        ("Backtracking Hill Climbing", backtracking_hill_climbing,     [5000]),
        ("Jumping Hill Climbing",      jumping_hill_climbing,          [jump_fn, 5000]),
        ("Random Restart HC",          random_restart_hill_climbing,   [restart_fn, 10, 500]),
    ]

    results = []
    print("📊 Đang chạy 7 thuật toán để thu thập dữ liệu vẽ biểu đồ...\n")
    for name, fn, extra in ALGOS:
        print(f"  → {name} ... ", end="", flush=True)
        t0 = time.time()
        best, path = fn(ini, get_nb, get_h, *extra)
        ms = (time.time() - t0) * 1000
        h_hist = _trace_h_history(ini, path, adapter)
        h_end  = get_h(best)
        print(f"{ms:.1f} ms | H(end)={h_end} | {len(path)} bước")
        results.append({
            "name":    name,
            "time_ms": round(ms, 2),
            "steps":   len(path),
            "h_start": h_hist[0],
            "h_end":   h_end,
            "h_history": h_hist,
            "success": h_end == 0,
        })
    return results


def collect_from_excel(path=EXCEL_INPUT):
    """
    Đọc dữ liệu từ file Excel do evaluate_algorithms.py tạo ra.
    Trả về list dict kết quả tương thích với hàm vẽ biểu đồ.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    results = []
    # Bảng có 7 cột nhóm (mỗi nhóm 3 cột: Bước, HĐ, Heuristic), bắt đầu từ cột 1
    for c_idx in range(7):
        base = c_idx * 3 + 1   # cột gốc của nhóm này (1-indexed)

        # Dòng 1: "Thuật toán: <name>"
        name_cell = ws.cell(row=1, column=base + 1).value or ""
        name = name_cell.replace("Thuật toán:", "").strip()

        # Dòng 2: "Thời gian: <ms> ms"
        time_cell = ws.cell(row=2, column=base + 1).value or ""
        try:
            ms = float(str(time_cell).replace("Thời gian:", "").replace("ms", "").strip())
        except ValueError:
            ms = 0.0

        # Dòng 4 trở đi: h_history ở cột base+2 (cột Heuristic)
        h_hist = []
        row = 4
        while True:
            h_val = ws.cell(row=row, column=base + 2).value
            if h_val is None or h_val == "":
                break
            try:
                h_hist.append(float(h_val))
            except (TypeError, ValueError):
                break
            row += 1
            if row > 10000:   # giới hạn an toàn
                break

        h_start = h_hist[0] if h_hist else 0
        h_end   = h_hist[-1] if h_hist else 0
        steps   = max(0, len(h_hist) - 1)

        results.append({
            "name":    name if name else f"Algo {c_idx+1}",
            "time_ms": ms,
            "steps":   steps,
            "h_start": h_start,
            "h_end":   h_end,
            "h_history": h_hist,
            "success": h_end == 0,
        })

    return results


# ══════════════════════════════════════════════════════════════════════════════
# 2.  VẼ BIỂU ĐỒ
# ══════════════════════════════════════════════════════════════════════════════

def _algo_color(name):
    return ALGO_COLORS.get(name, "#888888")


def plot_charts(results, save_path=OUTPUT_PNG):
    names    = [r["name"]    for r in results]
    times    = [r["time_ms"] for r in results]
    steps    = [r["steps"]   for r in results]
    h_starts = [r["h_start"] for r in results]
    h_ends   = [r["h_end"]   for r in results]
    success  = [r["success"] for r in results]
    colors   = [_algo_color(n) for n in names]

    # ── Thiết lập layout 3×2 + 1 biểu đồ lớn bên dưới ──────────────────────
    fig = plt.figure(figsize=(22, 26), facecolor="#0F1117")
    fig.suptitle(
        "ĐÁNH GIÁ HIỆU SUẤT 7 THUẬT TOÁN HILL CLIMBING – SOKOBAN AI",
        fontsize=18, fontweight="bold", color="#EAEAEA", y=0.98
    )

    # Lưới: 3 hàng × 2 cột cho 6 biểu đồ nhỏ + 1 hàng full-width ở dưới
    gs = fig.add_gridspec(
        4, 2,
        hspace=0.52, wspace=0.35,
        left=0.07, right=0.97,
        top=0.94, bottom=0.06
    )

    ax_time   = fig.add_subplot(gs[0, 0])   # 1. Tốc độ (ms)
    ax_steps  = fig.add_subplot(gs[0, 1])   # 2. Số bước
    ax_hbar   = fig.add_subplot(gs[1, 0])   # 3. Cải thiện Heuristic (bar)
    ax_pie    = fig.add_subplot(gs[1, 1])   # 4. Tỉ lệ thành công
    ax_scatter= fig.add_subplot(gs[2, 0])   # 5. Scatter: tốc độ vs bước
    ax_radar  = fig.add_subplot(gs[2, 1], polar=True)  # 6. Radar chart
    ax_conv   = fig.add_subplot(gs[3, :])   # 7. Đường hội tụ Heuristic

    _style_axes = lambda ax: (
        ax.set_facecolor("#1A1D27"),
        ax.tick_params(colors="#BBBBBB", labelsize=9),
        [s.set_color("#333344") for s in ax.spines.values()],
        ax.grid(color="#2A2D3E", linewidth=0.6, linestyle="--"),
    )

    short_names = [n.replace(" Hill Climbing", "\nHC").replace(" HC", "\nHC") for n in names]

    # ── 1. Biểu đồ cột: Tốc độ (ms) ────────────────────────────────────────
    _style_axes(ax_time)
    bars = ax_time.bar(range(len(names)), times, color=colors, edgecolor="#0F1117", linewidth=0.8, width=0.6)
    ax_time.set_xticks(range(len(names)))
    ax_time.set_xticklabels(short_names, fontsize=8, color="#CCCCCC")
    ax_time.set_ylabel("Thời gian (ms)", color="#CCCCCC", fontsize=10)
    ax_time.set_title("Thời Gian Thực Thi", color="#EAEAEA", fontsize=12, pad=10)
    for bar, t in zip(bars, times):
        ax_time.text(
            bar.get_x() + bar.get_width() / 2, bar.get_height() + max(times) * 0.01,
            f"{t:.1f}", ha="center", va="bottom", color="#FFFFFF", fontsize=8, fontweight="bold"
        )
    ax_time.set_ylim(0, max(times) * 1.2 if times else 1)

    # ── 2. Biểu đồ cột: Số bước ─────────────────────────────────────────────
    _style_axes(ax_steps)
    bars2 = ax_steps.bar(range(len(names)), steps, color=colors, edgecolor="#0F1117", linewidth=0.8, width=0.6)
    ax_steps.set_xticks(range(len(names)))
    ax_steps.set_xticklabels(short_names, fontsize=8, color="#CCCCCC")
    ax_steps.set_ylabel("Số bước", color="#CCCCCC", fontsize=10)
    ax_steps.set_title("Số Bước Đi", color="#EAEAEA", fontsize=12, pad=10)
    for bar, s in zip(bars2, steps):
        ax_steps.text(
            bar.get_x() + bar.get_width() / 2, bar.get_height() + max(steps) * 0.01,
            str(s), ha="center", va="bottom", color="#FFFFFF", fontsize=8, fontweight="bold"
        )
    ax_steps.set_ylim(0, max(steps) * 1.2 if steps else 1)

    # ── 3. Cải thiện Heuristic ───────────────────────────────────────────────
    _style_axes(ax_hbar)
    improvements = [hs - he for hs, he in zip(h_starts, h_ends)]
    bars3 = ax_hbar.barh(range(len(names)), improvements, color=colors, edgecolor="#0F1117", linewidth=0.8, height=0.5)
    ax_hbar.set_yticks(range(len(names)))
    ax_hbar.set_yticklabels(names, fontsize=8, color="#CCCCCC")
    ax_hbar.set_xlabel("Cải thiện H (H_start − H_end)", color="#CCCCCC", fontsize=10)
    ax_hbar.set_title("Mức Cải Thiện Heuristic", color="#EAEAEA", fontsize=12, pad=10)
    for bar, imp in zip(bars3, improvements):
        ax_hbar.text(
            bar.get_width() + max(improvements) * 0.01 if max(improvements) > 0 else 0.1,
            bar.get_y() + bar.get_height() / 2,
            f"{imp:.0f}", va="center", color="#FFFFFF", fontsize=8, fontweight="bold"
        )

    # ── 4. Tỉ lệ thành công (Pie) ────────────────────────────────────────────
    ax_pie.set_facecolor("#1A1D27")
    n_success = sum(success)
    n_fail    = len(success) - n_success
    pie_vals  = [n_success, n_fail] if n_fail > 0 else [n_success]
    pie_lbls  = [f"Thành công\n({n_success})", f"Thất bại\n({n_fail})"] if n_fail > 0 else [f"Thành công\n({n_success})"]
    pie_clrs  = ["#5BAD72", "#D13B3B"] if n_fail > 0 else ["#5BAD72"]
    wedges, texts, autotexts = ax_pie.pie(
        pie_vals, labels=pie_lbls, colors=pie_clrs,
        autopct="%1.0f%%", startangle=140,
        textprops={"color": "#FFFFFF", "fontsize": 10},
        wedgeprops={"edgecolor": "#0F1117", "linewidth": 1.5}
    )
    for at in autotexts:
        at.set_color("#0F1117")
        at.set_fontweight("bold")
    ax_pie.set_title("Tỉ Lệ Thành Công", color="#EAEAEA", fontsize=12, pad=14)

    # ── 5. Scatter: Tốc độ vs Số bước ────────────────────────────────────────
    _style_axes(ax_scatter)
    scatter_texts = []
    for i, r in enumerate(results):
        marker = "o" if r["success"] else "x"
        ax_scatter.scatter(r["time_ms"], r["steps"], color=_algo_color(r["name"]),
                           s=180, marker=marker, zorder=3, edgecolors="#FFFFFF", linewidths=0.8)
        txt = ax_scatter.text(r["time_ms"], r["steps"], r["name"].split()[0], 
                              fontsize=8, color="#DDDDDD", fontweight="bold")
        scatter_texts.append(txt)
        
    if scatter_texts:
        adjust_text(scatter_texts, ax=ax_scatter, arrowprops=dict(arrowstyle="-", color="#888888", lw=0.5))

    ax_scatter.set_xlabel("Thời gian thực thi (ms)", color="#CCCCCC", fontsize=10)
    ax_scatter.set_ylabel("Số bước đi", color="#CCCCCC", fontsize=10)
    ax_scatter.set_title("Tốc Độ vs Số Bước", color="#EAEAEA", fontsize=12, pad=10)

    # ── 6. Radar Chart: đánh giá đa chiều ─────────────────────────────────
    categories = ["Tốc độ\n(Nhanh)", "Số bước\n(Ít)", "Cải thiện \nHeuristic", "Tỉ lệ \nthành công"]
    N = len(categories)
    angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
    angles += angles[:1]

    ax_radar.set_facecolor("#1A1D27")
    ax_radar.set_theta_offset(np.pi / 2)
    ax_radar.set_theta_direction(-1)
    ax_radar.set_xticks(angles[:-1])
    ax_radar.set_xticklabels(categories, color="#CCCCCC", fontsize=10)
    ax_radar.tick_params(axis='x', pad=30)
    ax_radar.set_ylim(0, 1)
    ax_radar.yaxis.set_tick_params(labelcolor="#555566")
    ax_radar.grid(color="#2A2D3E", linewidth=0.7)
    ax_radar.set_title("Đánh Giá Đa Chiều", color="#EAEAEA", fontsize=12, pad=20)

    # Chuẩn hóa → [0, 1]
    def _norm(arr):
        mn, mx = min(arr), max(arr)
        if mx == mn: return [0.5] * len(arr)
        return [(x - mn) / (mx - mn) for x in arr]

    norm_time  = [1 - v for v in _norm(times)]   # nghịch: nhanh → cao
    norm_steps = [1 - v for v in _norm(steps)]   # nghịch: ít bước → cao
    norm_impr  = _norm(improvements)
    norm_succ  = [float(s) for s in success]

    for i, r in enumerate(results):
        vals = [norm_time[i], norm_steps[i], norm_impr[i], norm_succ[i]]
        vals += vals[:1]
        ax_radar.plot(angles, vals, color=_algo_color(r["name"]), linewidth=1.6, alpha=0.85)
        ax_radar.fill(angles, vals, color=_algo_color(r["name"]), alpha=0.15)

    # ── 7. Đường hội tụ Heuristic ─────────────────────────────────────────
    _style_axes(ax_conv)
    max_len = max(len(r["h_history"]) for r in results)
    for r in results:
        h_hist = r["h_history"]
        x = np.linspace(0, 100, len(h_hist))  # normalised 0-100 %
        ax_conv.plot(x, h_hist, color=_algo_color(r["name"]),
                     linewidth=2.0, alpha=0.9, label=r["name"])
    ax_conv.set_xlabel("Tiến trình tìm kiếm (%)", color="#CCCCCC", fontsize=11)
    ax_conv.set_ylabel("Giá trị Heuristic H", color="#CCCCCC", fontsize=11)
    ax_conv.set_title("Đường Hội Tụ Heuristic Theo Thời Gian", color="#EAEAEA", fontsize=13, pad=12)

    # ── Chú thích chung ────────────────────────────────────────────────────
    legend_patches = [
        mpatches.Patch(color=_algo_color(n), label=n)
        for n in names
    ]
    fig.legend(
        handles=legend_patches, loc="lower center",
        ncol=4, fontsize=9,
        facecolor="#1A1D27", edgecolor="#333344", labelcolor="#DDDDDD",
        bbox_to_anchor=(0.5, 0.01)
    )

    plt.savefig(save_path, dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    print(f"\n✅ Đã lưu Dashboard tổng → {save_path}")
    plt.close(fig) # Đóng fig thay vì show(block)

def save_individual_charts(results, out_dir="charts_output"):
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)

    names    = [r["name"]    for r in results]
    times    = [r["time_ms"] for r in results]
    steps    = [r["steps"]   for r in results]
    h_starts = [r["h_start"] for r in results]
    h_ends   = [r["h_end"]   for r in results]
    success  = [r["success"] for r in results]
    colors   = [_algo_color(n) for n in names]
    short_names = [n.replace(" Hill Climbing", "\nHC").replace(" HC", "\nHC") for n in names]

    def _prep_fig(figsize=(10, 6)):
        fig, ax = plt.subplots(figsize=figsize, facecolor="#0F1117")
        ax.set_facecolor("#1A1D27")
        ax.tick_params(colors="#BBBBBB", labelsize=10)
        for s in ax.spines.values():
            s.set_color("#333344")
        return fig, ax

    def _grid(ax):
        ax.grid(color="#2A2D3E", linewidth=0.6, linestyle="--")

    # 1. Tốc độ
    fig, ax = _prep_fig()
    _grid(ax)
    bars = ax.bar(range(len(names)), times, color=colors, edgecolor="#0F1117", width=0.6)
    ax.set_xticks(range(len(names)))
    ax.set_xticklabels(short_names, fontsize=9, color="#CCCCCC")
    ax.set_ylabel("Thời gian (ms)", color="#CCCCCC", fontsize=11)
    ax.set_title("Thời Gian Thực Thi", color="#EAEAEA", fontsize=14, pad=15)
    for bar, t in zip(bars, times):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(times)*0.01, f"{t:.1f}", ha="center", va="bottom", color="#FFFFFF", fontweight="bold")
    ax.set_ylim(0, max(times) * 1.2 if times else 1)
    fig.savefig(os.path.join(out_dir, "1_ThoiGian.png"), dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)

    # 2. Bước đi
    fig, ax = _prep_fig()
    _grid(ax)
    bars = ax.bar(range(len(names)), steps, color=colors, edgecolor="#0F1117", width=0.6)
    ax.set_xticks(range(len(names)))
    ax.set_xticklabels(short_names, fontsize=9, color="#CCCCCC")
    ax.set_ylabel("Số bước", color="#CCCCCC", fontsize=11)
    ax.set_title("Số Bước Đi", color="#EAEAEA", fontsize=14, pad=15)
    for bar, s in zip(bars, steps):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(steps)*0.01, str(s), ha="center", va="bottom", color="#FFFFFF", fontweight="bold")
    ax.set_ylim(0, max(steps) * 1.2 if steps else 1)
    fig.savefig(os.path.join(out_dir, "2_SoBuoc.png"), dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)

    # 3. Heuristic
    fig, ax = _prep_fig()
    _grid(ax)
    improvements = [hs - he for hs, he in zip(h_starts, h_ends)]
    bars = ax.barh(range(len(names)), improvements, color=colors, edgecolor="#0F1117", height=0.5)
    ax.set_yticks(range(len(names)))
    ax.set_yticklabels(names, fontsize=9, color="#CCCCCC")
    ax.set_xlabel("Cải thiện H (H_start − H_end)", color="#CCCCCC", fontsize=11)
    ax.set_title("Mức Cải Thiện Heuristic", color="#EAEAEA", fontsize=14, pad=15)
    for bar, imp in zip(bars, improvements):
        ax.text(bar.get_width() + max(improvements)*0.01 if max(improvements)>0 else 0.1, bar.get_y() + bar.get_height()/2, f"{imp:.0f}", va="center", color="#FFFFFF", fontweight="bold")
    fig.savefig(os.path.join(out_dir, "3_CaiThienHeuristic.png"), dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)

    # 4. Pie
    fig, ax = plt.subplots(figsize=(8, 8), facecolor="#0F1117")
    ax.set_facecolor("#1A1D27")
    n_success = sum(success)
    n_fail = len(success) - n_success
    pie_vals = [n_success, n_fail] if n_fail > 0 else [n_success]
    pie_lbls = [f"Thành công\n({n_success})", f"Thất bại\n({n_fail})"] if n_fail > 0 else [f"Thành công\n({n_success})"]
    pie_clrs = ["#5BAD72", "#D13B3B"] if n_fail > 0 else ["#5BAD72"]
    wedges, texts, autotexts = ax.pie(pie_vals, labels=pie_lbls, colors=pie_clrs, autopct="%1.0f%%", startangle=140, textprops={"color": "#FFFFFF", "fontsize": 12}, wedgeprops={"edgecolor": "#0F1117", "linewidth": 1.5})
    for at in autotexts:
        at.set_color("#0F1117")
        at.set_fontweight("bold")
    ax.set_title("Tỉ Lệ Thành Công", color="#EAEAEA", fontsize=14, pad=15)
    fig.savefig(os.path.join(out_dir, "4_TiLeThanhCong.png"), dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)

    # 5. Scatter
    fig, ax = _prep_fig()
    _grid(ax)
    indiv_scatter_texts = []
    for i, r in enumerate(results):
        marker = "o" if r["success"] else "x"
        ax.scatter(r["time_ms"], r["steps"], color=_algo_color(r["name"]), s=180, marker=marker, zorder=3, edgecolors="#FFFFFF", linewidths=0.8)
        txt = ax.text(r["time_ms"], r["steps"], r["name"].split()[0], fontsize=10, color="#DDDDDD", fontweight="bold")
        indiv_scatter_texts.append(txt)
        
    if indiv_scatter_texts:
        adjust_text(indiv_scatter_texts, ax=ax, arrowprops=dict(arrowstyle="-", color="#888888", lw=0.5))

    ax.set_xlabel("Thời gian thực thi (ms)", color="#CCCCCC", fontsize=11)
    ax.set_ylabel("Số bước đi", color="#CCCCCC", fontsize=11)
    ax.set_title("Tốc Độ vs Số Bước", color="#EAEAEA", fontsize=14, pad=15)
    fig.savefig(os.path.join(out_dir, "5_Scatter.png"), dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)

    # 6. Radar
    fig, ax = plt.subplots(figsize=(8, 8), facecolor="#0F1117", subplot_kw=dict(polar=True))
    ax.set_facecolor("#1A1D27")
    categories = ["Tốc độ\n(Nhanh)", "Số bước\n(Ít)", "Cải thiện \nHeuristic", "Tỉ lệ \nthành công"]
    N = len(categories)
    angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
    angles += angles[:1]
    ax.set_theta_offset(np.pi / 2)
    ax.set_theta_direction(-1)
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(categories, color="#CCCCCC", fontsize=11)
    ax.tick_params(axis='x', pad=30)
    ax.set_ylim(0, 1)
    ax.yaxis.set_tick_params(labelcolor="#555566")
    ax.grid(color="#2A2D3E", linewidth=0.7)
    ax.set_title("Đánh Giá Đa Chiều", color="#EAEAEA", fontsize=16, pad=25)
    
    def _norm(arr):
        mn, mx = min(arr), max(arr)
        if mx == mn: return [0.5] * len(arr)
        return [(x - mn) / (mx - mn) for x in arr]
    norm_time  = [1 - v for v in _norm(times)]
    norm_steps = [1 - v for v in _norm(steps)]
    norm_impr  = _norm([hs - he for hs, he in zip(h_starts, h_ends)])
    norm_succ  = [float(s) for s in success]
    
    for i, r in enumerate(results):
        vals = [norm_time[i], norm_steps[i], norm_impr[i], norm_succ[i]]
        vals += vals[:1]
        ax.plot(angles, vals, color=_algo_color(r["name"]), linewidth=1.6, alpha=0.85)
        ax.fill(angles, vals, color=_algo_color(r["name"]), alpha=0.15)
    fig.savefig(os.path.join(out_dir, "6_Radar.png"), dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)

    # 7. Line
    fig, ax = plt.subplots(figsize=(14, 6), facecolor="#0F1117")
    ax.set_facecolor("#1A1D27")
    ax.tick_params(colors="#BBBBBB", labelsize=10)
    for s in ax.spines.values():
        s.set_color("#333344")
    _grid(ax)
    
    for r in results:
        h_hist = r["h_history"]
        x = np.linspace(0, 100, len(h_hist))
        ax.plot(x, h_hist, color=_algo_color(r["name"]), linewidth=2.0, alpha=0.9, label=r["name"])
    ax.set_xlabel("Tiến trình tìm kiếm (%)", color="#CCCCCC", fontsize=12)
    ax.set_ylabel("Giá trị Heuristic H", color="#CCCCCC", fontsize=12)
    ax.set_title("Đường Hội Tụ Heuristic Theo Thời Gian", color="#EAEAEA", fontsize=14, pad=15)
    ax.legend(loc="upper right", fontsize=10, facecolor="#1A1D27", edgecolor="#333344", labelcolor="#DDDDDD", ncol=3)
    fig.savefig(os.path.join(out_dir, "7_LineChart.png"), dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    print(f"✅ Đã lưu 7 biểu đồ riêng biệt vào thư mục '{out_dir}'")



# ══════════════════════════════════════════════════════════════════════════════
# 3.  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Vẽ biểu đồ đánh giá thuật toán Sokoban AI")
    parser.add_argument(
        "--excel", action="store_true",
        help="Đọc dữ liệu từ file Excel đã có thay vì chạy lại thuật toán"
    )
    parser.add_argument(
        "--file", default=EXCEL_INPUT,
        help=f"Đường dẫn file Excel đầu vào (mặc định: {EXCEL_INPUT})"
    )
    parser.add_argument(
        "--out", default=OUTPUT_PNG,
        help=f"Đường dẫn file PNG kết quả (mặc định: {OUTPUT_PNG})"
    )
    args = parser.parse_args()

    if args.excel:
        if not os.path.exists(args.file):
            print(f"❌ Không tìm thấy file: {args.file}")
            print("   Hãy chạy 'python evaluate_algorithms.py' trước để tạo file Excel,")
            print("   hoặc chạy 'python chart_analysis.py' (không có --excel) để tự thu thập dữ liệu.")
            sys.exit(1)
        print(f"📂 Đang đọc dữ liệu từ: {args.file}")
        results = collect_from_excel(args.file)
    else:
        results = collect_from_run()

    print(f"\n📊 Đang vẽ {len(results)} thuật toán...")
    plot_charts(results, save_path=args.out)
    save_individual_charts(results)
    
    # Hiển thị ảnh tổng ở cuối cùng nếu cần thiết hoặc chỉ lưu file
    # plt.show() được gỡ ra để tiến trình không bị chặn khi chạy tự động


if __name__ == "__main__":
    main()
